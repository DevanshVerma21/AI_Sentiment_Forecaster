"""
Report Generator CLI
====================
Generates PDF and Excel reports from analytics data.

Usage:
    python -m reports.generate --period weekly
    python -m reports.generate --period monthly --format pdf
    python -m reports.generate --product "electronics"
"""
import os
import sys
import json
import logging
import argparse
from datetime import datetime

import pandas as pd

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
sys.path.insert(0, os.path.join(PROJECT_ROOT, "backend"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("report_generator")

ANALYTICS_DIR = os.path.join(PROJECT_ROOT, "data", "analytics")
ENRICHED_DIR = os.path.join(PROJECT_ROOT, "data", "enriched")
REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports", "output")


def _load_analytics() -> dict:
    """Load all analytics data."""
    data = {}

    # Summary stats
    summary_path = os.path.join(ANALYTICS_DIR, "summary_stats.json")
    if os.path.exists(summary_path):
        with open(summary_path) as f:
            data["summary"] = json.load(f)

    # Product comparison
    product_path = os.path.join(ANALYTICS_DIR, "product_comparison.parquet")
    if os.path.exists(product_path):
        data["products"] = pd.read_parquet(product_path)

    # Topic trends
    topic_path = os.path.join(ANALYTICS_DIR, "topic_trends.parquet")
    if os.path.exists(topic_path):
        data["topics"] = pd.read_parquet(topic_path)

    # Daily sentiment trends
    trends_path = os.path.join(ANALYTICS_DIR, "sentiment_trends.parquet")
    if os.path.exists(trends_path):
        data["trends"] = pd.read_parquet(trends_path)

    # Alerts log
    alerts_path = os.path.join(PROJECT_ROOT, "logs", "alerts.log")
    if os.path.exists(alerts_path):
        alerts = []
        with open(alerts_path) as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        alerts.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
        data["alerts"] = alerts[-20:]  # last 20 alerts

    return data


def generate_pdf_report(period: str, product: str = None) -> str:
    """Generate a PDF summary report using ReportLab."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors

    os.makedirs(REPORTS_DIR, exist_ok=True)
    data = _load_analytics()

    if not data.get("summary"):
        logger.error("No analytics data found. Run: python -m pipelines.analytics_pipeline")
        return ""

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{date_str}_{period}_report.pdf"
    if product:
        filename = f"{date_str}_{period}_{product.replace(' ', '_')}_report.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)

    doc = SimpleDocTemplate(filepath, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"],
        fontSize=22, textColor=colors.HexColor("#0066CC"),
        spaceAfter=20, alignment=1,
    )
    heading_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading2"],
        fontSize=14, textColor=colors.HexColor("#0066CC"),
        spaceAfter=10, spaceBefore=12,
    )

    summary = data["summary"]

    # ── Title Page ──
    story.append(Paragraph("AI Market Trend & Sentiment Report", title_style))
    story.append(Paragraph(f"Period: {period.title()} | Generated: {date_str}", styles["Normal"]))
    if product:
        story.append(Paragraph(f"Product Focus: {product}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * inch))

    # ── Executive Summary ──
    story.append(Paragraph("Executive Summary", heading_style))
    story.append(Paragraph(
        f"This report covers {summary['total_records']} data points across "
        f"{summary['unique_products']} products/topics from {summary['unique_sources']} sources. "
        f"Overall sentiment: {summary['positive_pct']}% positive, {summary['negative_pct']}% negative.",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.2 * inch))

    # ── Key Metrics Table ──
    story.append(Paragraph("Key Metrics", heading_style))
    metrics_data = [
        ["Metric", "Value"],
        ["Total Records", str(summary["total_records"])],
        ["Mean Sentiment", f"{summary['mean_sentiment_score']:.3f}"],
        ["Positive %", f"{summary['positive_pct']}%"],
        ["Negative %", f"{summary['negative_pct']}%"],
        ["Unique Products", str(summary["unique_products"])],
        ["Unique Topics", str(summary["unique_topics"])],
    ]
    t = Table(metrics_data, colWidths=[2.5 * inch, 2.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F0F8FF")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))

    # ── Top Products ──
    if "products" in data:
        story.append(PageBreak())
        story.append(Paragraph("Top Products / Topics by Sentiment", heading_style))

        products_df = data["products"].sort_values("mean_score", ascending=False).head(10)
        prod_data = [["Product", "Avg Score", "Count", "Positive %", "Negative %"]]
        for _, row in products_df.iterrows():
            prod_data.append([
                str(row["product_or_topic"])[:30],
                f"{row['mean_score']:.3f}",
                str(int(row["count"])),
                f"{row.get('pos_pct', 0):.0f}%",
                f"{row.get('neg_pct', 0):.0f}%",
            ])

        t = Table(prod_data, colWidths=[2 * inch, 1 * inch, 0.8 * inch, 1 * inch, 1 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t)

    # ── Top Topics ──
    if "topics" in data:
        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("Emerging Topics", heading_style))

        topics_df = data["topics"].head(10)
        topic_data = [["Topic", "Mentions", "Avg Sentiment"]]
        for _, row in topics_df.iterrows():
            topic_data.append([
                str(row["topic_label"])[:40],
                str(int(row["count"])),
                f"{row['mean_score']:.3f}",
            ])

        t = Table(topic_data, colWidths=[3 * inch, 1.2 * inch, 1.5 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t)

    # ── Recent Alerts ──
    if data.get("alerts"):
        story.append(PageBreak())
        story.append(Paragraph("Recent Alerts", heading_style))
        for alert in data["alerts"][-5:]:
            severity = alert.get("severity", "INFO")
            color = {"CRITICAL": "red", "WARNING": "orange"}.get(severity, "black")
            story.append(Paragraph(
                f'<font color="{color}">[{severity}]</font> {alert.get("message", "")}',
                styles["Normal"],
            ))
            story.append(Spacer(1, 0.05 * inch))

    doc.build(story)
    logger.info(f"PDF report generated: {filepath}")
    return filepath


def generate_excel_report(period: str, product: str = None) -> str:
    """Generate an Excel report with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    os.makedirs(REPORTS_DIR, exist_ok=True)
    data = _load_analytics()

    if not data.get("summary"):
        logger.error("No analytics data found.")
        return ""

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{date_str}_{period}_report.xlsx"
    if product:
        filename = f"{date_str}_{period}_{product.replace(' ', '_')}_report.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Sentiment Report — {period.title()}"
    ws["A1"].font = Font(bold=True, size=14, color="0066CC")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Generated: {date_str}"

    summary = data["summary"]
    row = 4
    for key, value in summary.items():
        ws[f"A{row}"] = key.replace("_", " ").title()
        ws[f"B{row}"] = str(value)
        row += 1

    # ── Products Sheet ──
    if "products" in data:
        ws2 = wb.create_sheet("Products")
        products_df = data["products"].sort_values("mean_score", ascending=False)
        headers = ["Product", "Count", "Mean Score", "Std", "Positive %", "Negative %"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, (_, row_data) in enumerate(products_df.iterrows(), 2):
            ws2.cell(row=r, column=1, value=str(row_data["product_or_topic"]))
            ws2.cell(row=r, column=2, value=int(row_data["count"]))
            ws2.cell(row=r, column=3, value=round(row_data["mean_score"], 4))
            ws2.cell(row=r, column=4, value=round(row_data.get("std_score", 0), 4))
            ws2.cell(row=r, column=5, value=round(row_data.get("pos_pct", 0), 1))
            ws2.cell(row=r, column=6, value=round(row_data.get("neg_pct", 0), 1))

        for col in range(1, 7):
            ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Topics Sheet ──
    if "topics" in data:
        ws3 = wb.create_sheet("Topics")
        topics_df = data["topics"]
        headers = ["Topic ID", "Topic Label", "Count", "Mean Sentiment"]
        for col, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, (_, row_data) in enumerate(topics_df.iterrows(), 2):
            ws3.cell(row=r, column=1, value=int(row_data["topic_id"]))
            ws3.cell(row=r, column=2, value=str(row_data["topic_label"]))
            ws3.cell(row=r, column=3, value=int(row_data["count"]))
            ws3.cell(row=r, column=4, value=round(row_data["mean_score"], 4))

    # ── Alerts Sheet ──
    if data.get("alerts"):
        ws4 = wb.create_sheet("Alerts")
        headers = ["Severity", "Type", "Product", "Message", "Timestamp"]
        for col, h in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, alert in enumerate(data["alerts"], 2):
            ws4.cell(row=r, column=1, value=alert.get("severity", ""))
            ws4.cell(row=r, column=2, value=alert.get("type", ""))
            ws4.cell(row=r, column=3, value=alert.get("product", ""))
            ws4.cell(row=r, column=4, value=alert.get("message", ""))
            ws4.cell(row=r, column=5, value=alert.get("timestamp", ""))

        for col in range(1, 6):
            ws4.column_dimensions[get_column_letter(col)].width = 22

    wb.save(filepath)
    logger.info(f"Excel report generated: {filepath}")
    return filepath


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="TrendAI Report Generator")
    parser.add_argument("--period", choices=["daily", "weekly", "monthly"], default="weekly")
    parser.add_argument("--format", choices=["pdf", "xlsx", "both"], default="both")
    parser.add_argument("--product", type=str, default=None, help="Filter for specific product")
    args = parser.parse_args()

    generated = []

    if args.format in ("pdf", "both"):
        path = generate_pdf_report(args.period, product=args.product)
        if path:
            generated.append(path)

    if args.format in ("xlsx", "both"):
        path = generate_excel_report(args.period, product=args.product)
        if path:
            generated.append(path)

    if generated:
        print(f"\nGenerated reports:")
        for p in generated:
            print(f"  {p}")
    else:
        print("No reports generated. Ensure analytics data exists.")
