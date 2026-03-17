"""
Report Generation Service
Generates PDF and Excel reports with sentiment analysis and insights
"""
import logging
from typing import Dict, List, Any
from datetime import datetime
import json
import os

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate PDF and Excel reports for sentiment analysis"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Report generator initialized, output: {output_dir}")
    
    def generate_pdf_report(self,
                          product: str,
                          analysis_data: Dict[str, Any],
                          insights: Dict[str, Any] = None) -> str:
        """
        Generate PDF sentiment report
        
        Args:
            product: Product name
            analysis_data: Analysis results from realtime analyzer
            insights: Additional insights
        
        Returns:
            Path to generated PDF
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sentiment_report_{product.replace(' ', '_')}_{timestamp}.pdf"
            filepath = os.path.join(self.output_dir, filename)
            
            logger.info(f"Generating PDF report: {filepath}")
            
            # Create PDF
            doc = SimpleDocTemplate(filepath, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#0066CC'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#0066CC'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            # Title
            story.append(Paragraph(f"Sentiment Analysis Report", title_style))
            story.append(Paragraph(f"Product: {product}", styles['Heading2']))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", heading_style))
            summary_text = analysis_data.get("summary", "No summary available")
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            # Key Metrics
            story.append(Paragraph("Key Metrics", heading_style))
            metrics_data = [
                ["Metric", "Value"],
                ["Articles Analyzed", str(analysis_data.get("article_count", 0))],
                ["Sentiment Score", f"{analysis_data.get('sentiment_score', 0):.2f}"],
                ["Data Source", analysis_data.get("source", "N/A")],
                ["Generated At", analysis_data.get("generated_at", "N/A")],
            ]
            
            metrics_table = Table(metrics_data, colWidths=[2.5*inch, 2.5*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metrics_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Sentiment Breakdown
            story.append(Paragraph("Sentiment Breakdown", heading_style))
            breakdown = analysis_data.get("sentiment_breakdown", {})
            breakdown_data = [
                ["Sentiment", "Count", "Percentage"],
                ["Positive", str(breakdown.get("positive", 0)), f"{(breakdown.get('positive', 0) / (analysis_data.get('article_count', 1)) * 100):.1f}%"],
                ["Neutral", str(breakdown.get("neutral", 0)), f"{(breakdown.get('neutral', 0) / (analysis_data.get('article_count', 1)) * 100):.1f}%"],
                ["Negative", str(breakdown.get("negative", 0)), f"{(breakdown.get('negative', 0) / (analysis_data.get('article_count', 1)) * 100):.1f}%"],
            ]
            
            breakdown_table = Table(breakdown_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch])
            breakdown_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(breakdown_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Top Articles
            story.append(PageBreak())
            story.append(Paragraph("Top Articles", heading_style))
            
            top_articles = analysis_data.get("top_articles", [])[:5]
            for idx, article in enumerate(top_articles, 1):
                story.append(Paragraph(f"{idx}. {article.get('title', 'No title')[:60]}...", styles['Heading3']))
                story.append(Paragraph(
                    f"Platform: {article.get('platform', 'N/A')} | "
                    f"Sentiment: {article.get('sentiment_label', 'N/A')} | "
                    f"Confidence: {article.get('confidence_score', 0):.2f}",
                    styles['Normal']
                ))
                story.append(Paragraph(
                    article.get('description', 'No description')[:200] + "...",
                    styles['Normal']
                ))
                story.append(Spacer(1, 0.1*inch))
            
            # Insights
            story.append(PageBreak())
            story.append(Paragraph("Insights & Recommendations", heading_style))
            
            insights_list = analysis_data.get("insights", [])
            for insight in insights_list[:5]:
                story.append(Paragraph(f"• {insight}", styles['Normal']))
                story.append(Spacer(1, 0.1*inch))
            
            # Build PDF
            doc.build(story)
            logger.info(f"PDF report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"PDF generation error: {str(e)}")
            return None
    
    def generate_excel_report(self,
                            product: str,
                            analysis_data: Dict[str, Any],
                            articles: List[Dict[str, Any]] = None) -> str:
        """
        Generate Excel report with detailed metrics
        
        Args:
            product: Product name
            analysis_data: Analysis results
            articles: List of analyzed articles
        
        Returns:
            Path to generated Excel file
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sentiment_report_{product.replace(' ', '_')}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            logger.info(f"Generating Excel report: {filepath}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            
            # Styles
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title Section
            ws['A1'] = f"Sentiment Analysis Report: {product}"
            ws['A1'].font = Font(bold=True, size=14, color="0066CC")
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True, size=10)
            ws.merge_cells('A2:D2')
            
            # Summary Section
            row = 4
            ws[f'A{row}'] = "Summary Metrics"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            summary_metrics = [
                ("Total Articles", analysis_data.get("article_count", 0)),
                ("Sentiment Score", f"{analysis_data.get('sentiment_score', 0):.2f}"),
                ("Data Source", analysis_data.get("source", "N/A")),
                ("Positive Reviews", analysis_data.get("sentiment_breakdown", {}).get("positive", 0)),
                ("Neutral Reviews", analysis_data.get("sentiment_breakdown", {}).get("neutral", 0)),
                ("Negative Reviews", analysis_data.get("sentiment_breakdown", {}).get("negative", 0)),
            ]
            
            for label, value in summary_metrics:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
            
            # Articles Sheet
            if articles:
                ws_articles = wb.create_sheet("Articles")
                
                # Headers
                headers = ["Title", "Platform", "Sentiment", "Confidence", "Date", "Summary"]
                for col, header in enumerate(headers, 1):
                    cell = ws_articles.cell(row=1, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Articles
                for row, article in enumerate(articles[:100], 2):  # Limit to 100 articles
                    ws_articles.cell(row=row, column=1).value = article.get("title", "")[:50]
                    ws_articles.cell(row=row, column=2).value = article.get("platform", "")
                    ws_articles.cell(row=row, column=3).value = article.get("sentiment_label", "")
                    ws_articles.cell(row=row, column=4).value = article.get("confidence_score", 0)
                    ws_articles.cell(row=row, column=5).value = article.get("published_date", "")
                    ws_articles.cell(row=row, column=6).value = article.get("description", "")[:100]
                
                # Auto-size columns
                for col in range(1, 7):
                    ws_articles.column_dimensions[get_column_letter(col)].width = 15
            
            # Trends Sheet
            if "yearly_sentiment_trend" in analysis_data:
                ws_trends = wb.create_sheet("Trends")
                
                headers = ["Year", "Sentiment Score", "Sample Count"]
                for col, header in enumerate(headers, 1):
                    cell = ws_trends.cell(row=1, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                
                for row, trend in enumerate(analysis_data["yearly_sentiment_trend"], 2):
                    ws_trends.cell(row=row, column=1).value = trend.get("year")
                    ws_trends.cell(row=row, column=2).value = trend.get("score", 0)
                    ws_trends.cell(row=row, column=3).value = trend.get("samples", 0)
            
            # Save
            wb.save(filepath)
            logger.info(f"Excel report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Excel generation error: {str(e)}")
            return None
    
    def generate_batch_report(self, products_analysis: Dict[str, Dict[str, Any]]) -> str:
        """
        Generate comparison report across multiple products
        
        Args:
            products_analysis: Dictionary of {product_name: analysis_data}
        
        Returns:
            Path to generated Excel file
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"batch_sentiment_report_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            logger.info(f"Generating batch report: {filepath}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison"
            
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Headers
            headers = ["Product", "Sentiment Score", "Total Articles", "Positive %", "Negative %", "Neutral %", "Data Source"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            for row, (product, data) in enumerate(products_analysis.items(), 2):
                breakdown = data.get("sentiment_breakdown", {})
                total = data.get("article_count", 1)
                
                ws.cell(row=row, column=1).value = product
                ws.cell(row=row, column=2).value = f"{data.get('sentiment_score', 0):.2f}"
                ws.cell(row=row, column=3).value = total
                ws.cell(row=row, column=4).value = f"{(breakdown.get('positive', 0) / total * 100):.1f}%"
                ws.cell(row=row, column=5).value = f"{(breakdown.get('negative', 0) / total * 100):.1f}%"
                ws.cell(row=row, column=6).value = f"{(breakdown.get('neutral', 0) / total * 100):.1f}%"
                ws.cell(row=row, column=7).value = data.get("source", "N/A")
            
            # Auto-size
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = 18
            
            wb.save(filepath)
            logger.info(f"Batch report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Batch report generation error: {str(e)}")
            return None


# Global instance
_report_gen = None

def get_report_generator(output_dir: str = "reports") -> ReportGenerator:
    """Get or create report generator"""
    global _report_gen
    if _report_gen is None:
        _report_gen = ReportGenerator(output_dir)
    return _report_gen
